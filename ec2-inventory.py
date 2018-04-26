"""
!/usr/bin/env python
   Prerequisites:
       pip install openpyxl boto3
       price.json file generated by ec2-price-json-generator.py
   Author: Cosmin.Gavagiuc@gmail.com

   Usage : python inventory.py AWS-profile-name

   This script create a XLSX report with EC2 instances details report on owned by AWS account.
"""
import boto3
import os
import re
import time
from datetime import datetime, timedelta
import logging
from botocore.exceptions import ClientError
import tempfile
import yaml
import json
import csv
import sys
import string
from boto3.session import Session
from operator import itemgetter
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color

# static globals

os.environ['AWS_PROFILE'] = sys.argv[1]
os.environ['AWS_DEFAULT_REGION'] = "eu-west-1"
timestr = time.strftime("%Y%m%d-%H%M%S")
logger = logging.getLogger()
logger.setLevel(logging.INFO)

def monitor_cw(instance_id, region):
    # returns cpu utilization from cloudwatch
    now = datetime.utcnow()
    past = now - timedelta(minutes=10090)
    future = now + timedelta(minutes=10)
    cwclient = boto3.client('cloudwatch', region_name=region)
    results = cwclient.get_metric_statistics(
        Namespace='AWS/EC2',
        MetricName='CPUUtilization',
        Dimensions=[{'Name': 'InstanceId', 'Value': instance_id}],
        StartTime=past,
        EndTime=future,
        Period=86400,
        Statistics=['Average'])
    datapoints = results['Datapoints']
    load = ''
    if datapoints:
        last_datapoint = sorted(datapoints, key=itemgetter('Timestamp'))[-1]
        utilization = last_datapoint['Average']
        load = round((utilization), 2)
    return load

def monitor_ec2(region):
    # returns row with ec2 details.
    client = boto3.client('ec2', region_name=region)
    paginator = client.get_paginator('describe_instances')
    response_iterator = paginator.paginate()
    for page in response_iterator:
        for obj in page['Reservations']:
            for instance in obj['Instances']:
                InstanceName=None
                Platform = "linux"
                if instance['State']['Name'] != 'terminated':
                    for tag in instance["Tags"]:
                        if tag["Key"] == 'Name':InstanceName = tag["Value"]
                #get attached volumes
                PrivateIP = None
                PublicIPADDDR = None
                print(InstanceName)
                ec2 = boto3.resource('ec2', region_name=region)
                InstanceDetails = ec2.Instance(instance['InstanceId'])
                Volumes = InstanceDetails.volumes.all()
                ec2vol = list()
                for Volume in Volumes:
                    Vol = ec2.Volume(id=Volume.id)
                    ec2vol.append(Vol.attachments[0][u'Device'])
                    ec2vol.append(Vol.size)
                for inet in instance['NetworkInterfaces']:
                    if 'Association' in inet and 'PublicIp' in inet['Association']: PublicIPADDDR = inet['Association']['PublicIp']
                if 'Platform' in instance: Platform = instance['Platform']
                if 'PrivateIpAddress' in instance: PrivateIP = instance['PrivateIpAddress']
                HW = data['compute']['models'][region][instance["InstanceType"]]
                row = list()
                row.append(instance['Placement']['AvailabilityZone'])
                row.append(InstanceName)
                row.append(instance["InstanceId"])
                row.append(instance["InstanceType"])
                row.append(Platform)
                row.append(PublicIPADDDR)
                row.append(PrivateIP)
                row.append(instance['State']['Name'])
                row.append(instance['LaunchTime'])
                row.append(Account)
                row.append(HW['CPU'])
                row.append(monitor_cw(instance["InstanceId"], region))
                row.append(HW['ECU'])
                row.append(HW['memoryGiB'])
                row.extend(ec2vol)
                ws.append(row)

def get_regions():
    """
        List all AWS region available
        :return: list of regions
    """
    client = boto3.client('ec2')
    regions = [region['RegionName'] for region in client.describe_regions()['Regions']]
    return regions

def format_xlsx(ws):
    #
    #       Formating XLSX
    #   TOP Row
        for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
            for cell in rows:
                cell.fill = PatternFill("solid", fgColor="0066a1")
                cell.font = Font(color="00FFFFFF",bold=True)
    #   Format entire table
        thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    # set wrapText
        alignment=Alignment(wrap_text=True)
        for col in ws.columns:
            for cell in col:
                cell.alignment = alignment
    # set weight
        for col in ws.columns:
            max_length = 0
            column = col[0].column #Get the column name
            for cell in col[1:]:
                cell.border = thin_border
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                        if ' ' in str(cell.value): max_length = ((len(cell.value)+2)*1.4)/2
                except:
                    pass
            adjusted_width = (max_length + 3)
            ws.column_dimensions[column].width = adjusted_width

def init_moniroting():
    """
        Script initialisation
        :retur: None
    """
    # Setting regions
    global region_list
    region_list = get_regions()

if __name__ == '__main__':
    import sys
    logging.basicConfig(level=logging.WARNING)
    try:
        print('Loading price.json')
        with open('price.json') as json_file:
            data = json.load(json_file)
        wb = Workbook()
        # grab the active worksheet
        ws = wb.active
        ws.append(['Placement','Name','Instance ID','Instance Type','Platform','Public IP','Private IP','Instance State','LaunchTime','AWS Account','CPU','CPU Utilization Avg','ECU','memory GiB','Volume','Size GiB','Volume','Size GiB','Volume','Size GiB','Volume','Size GiB','Volume','Size GiB'])
        global Account
        iam = boto3.client("iam")
        paginator = iam.get_paginator('list_account_aliases')
        for response in paginator.paginate():
            Account = "\n".join(response['AccountAliases'])
            # Calling functions for every region
        init_moniroting()
        for region in region_list:
            monitor_ec2(region)
        ws=format_xlsx(ws)
        wb.save("inventory-"+Account+"-"+timestr+".xlsx")
    except ClientError as e:
        logger.error(e)
    except Exception as err:
        logger.error(err)
